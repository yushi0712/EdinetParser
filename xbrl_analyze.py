import os
import sys
import pandas as pd
import numpy as np
import xbrl_common
import time

from os import path
from xbrl_proc import read_xbrl
from statistics import mean, median,variance,stdev


list_focused_industries = ["精密機器", "機械", "電気機器"]


#---------------------------------------
# ASR Summaryファイルを読み込む
#---------------------------------------
print("◆XBRL Contentsファイルを読み込む", end="")
asr_summary_file = pd.ExcelFile(xbrl_common.XBRL_ROOT_PATH + "/" + xbrl_common.ASR_SUMMARY_FILE_NAME)
df_asr_summary = asr_summary_file.parse(sheet_name="OrgData")
df_industry = pd.DataFrame(list(set(df_asr_summary["業種"])), columns=["業種"])

list_column = ["従業員数", "純資産", "総資産", "売上高", "純利益", "株価収益率", "営業CF", "投資CF", "財務CF", "現金"]
list_year = ["", "(P1Y)", "(P2Y)", "(P3Y)", "(P4Y)"]
dict_column= dict()
for col in list_column:
    dict_column[col] = [col+s for s in list_year]
print("  -> 完了")

#---------------------------------------
# 提出社ごとの統計値を計算
#---------------------------------------
skip_flag = True
dict_statstic_index = dict()
if not skip_flag:
    print("◆提出社ごとの統計値を計算")
    for index, row in df_asr_summary.iterrows():
        # 財務CFと株価収益率以外の空欄をカウント
        null_num = row[6:].isnull().sum()
        cnt = 0
        for idx, itm in row.iteritems(): 
            if ("財務CF" in idx) or ("株価収益率" in idx):
                if pd.isnull(itm):
                    cnt += 1
        # 計算
        try:
            dict_statstic_index["売上／人員"] = [x / y for (x, y) in zip(row[dict_column["売上高"]], row[dict_column["従業員数"]])]
            dict_statstic_index["営業CF／人員"] = [x / y for (x, y) in zip(row[dict_column["営業CF"]], row[dict_column["従業員数"]])]
            dict_statstic_index["FCF／人員"] = [(x+y) / z for (x, y, z) in zip(row[dict_column["営業CF"]], row[dict_column["投資CF"]], row[dict_column["従業員数"]])]
            dict_statstic_index["純利益／人員"] = [x / y for (x, y) in zip(row[dict_column["純利益"]], row[dict_column["従業員数"]])]
            dict_statstic_index["自己資本比率"] = [x / y for (x, y) in zip(row[dict_column["純資産"]], row[dict_column["総資産"]])]
            dict_statstic_index["ROE"] = [x / y for (x, y) in zip(row[dict_column["純利益"]], row[dict_column["純資産"]])]
            dict_statstic_index["ROA"] = [x / y for (x, y) in zip(row[dict_column["純利益"]], row[dict_column["総資産"]])]
        except ZeroDivisionError:
            pass

        # 列追加        
        df_asr_summary.at[index, "空欄数"] = null_num - cnt
        for col in list_column:
            df_asr_summary.at[index, col+"(5年平均)"] = row[dict_column[col]].mean()
        for k in dict_statstic_index.keys():
            df_asr_summary.at[index, k+"(5年平均)"] = mean(dict_statstic_index[k])
        for k in dict_statstic_index.keys():
            for i, y in enumerate(list_year):
                df_asr_summary.at[index, k+y] = (dict_statstic_index[k])[i]
        print("\r    {0}/{1}".format(index+1, len(df_asr_summary)), end="")
                    
    df_asr_summary_available = df_asr_summary[df_asr_summary["空欄数"]==0]
    print("  -> 完了")


#---------------------------------------
# 業種ごとにDataFrameを作成
#---------------------------------------
df_specified_industries = dict()    
for index, row in df_industry.iterrows():
    q_word = "業種 == \"{0}\"".format(row["業種"])
    df_specified_industries[row["業種"]] = df_asr_summary_available.query(q_word)


#---------------------------------------
# 各業種ごとの統計値計算
#---------------------------------------
df_corr_mat= dict()    
for index, row in df_industry.iterrows():
    industry = row["業種"]
    df_temp = df_specified_industries[industry]
    df_available = df_temp[df_temp["空欄数"]==0]

    df_industry.at[index, "提出社数"] = len(df_temp)
    df_industry.at[index, "有効提出社数"] = len(df_available)
    df_industry.at[index, "有効提出社率"] = 100*df_industry.at[index, "有効提出社数"]/df_industry.at[index, "提出社数"]
    df_industry.at[index, "従業員数(平均)"] = int(df_available["従業員数"].mean())
    df_industry.at[index, "従業員数(中央)"] = int(df_available["従業員数"].median())
    df_industry.at[index, "売上高(平均)"] = int(df_available["売上高"].mean())
    df_industry.at[index, "売上高(中央)"] = int(df_available["売上高"].median())
    df_industry.at[index, "売上／人員(平均)"] = int(df_available["売上／人員"].mean())
    df_industry.at[index, "売上／人員(中央)"] = int(df_available["売上／人員"].median())
    df_industry.at[index, "営業CF／人員(平均)"] = int(df_available["営業CF／人員"].mean())
    df_industry.at[index, "営業CF／人員(中央)"] = int(df_available["営業CF／人員"].median())
    df_industry.at[index, "ROA(平均)"] = df_available["ROA"].mean()
    df_industry.at[index, "ROA(中央)"] = df_available["ROA"].median()

    df_industry.at[index, "人員x売上高(相関)"] = np.corrcoef(df_available["従業員数"], df_available["売上高"])[0, 1]#corr_mat.at["従業員数", "売上高"]
    df_industry.at[index, "人員x総資産(相関)"] = np.corrcoef(df_available["従業員数"], df_available["総資産"])[0, 1]
    df_industry.at[index, "売上高x総資産(相関)"] = np.corrcoef(df_available["売上高"], df_available["総資産"])[0, 1]

#---------------------------------------
# 注目業界
#---------------------------------------
df_asr_summary_focused = pd.DataFrame(columns=df_asr_summary_available.columns)
for fc in list_focused_industries:
    df_asr_summary_focused = pd.concat([df_asr_summary_focused, df_specified_industries[fc]])
#従業員数で範囲制限
#df_asr_summary_focused = df_asr_summary_focused[df_asr_summary_focused["従業員数"] < 5000]
df_asr_summary_focused = df_asr_summary_focused[df_asr_summary_focused["従業員数"] > 1000]

#---------------------------------------
# 分析結果をExcelファイルに保存
#---------------------------------------
print("◆分析結果をExcelファイルに保存", end="")
import openpyxl as px
wb = px.Workbook()
ws_asr_summary = wb.create_sheet(title="全提出社者")
ws = wb.create_sheet(title="Focused")

from openpyxl.styles.fonts import Font
font_colmun = Font(b=True, sz=9)
font_cell = Font(sz=9)
align_column = px.styles.Alignment(horizontal="center", vertical="center")
align_cell = px.styles.Alignment(vertical="center")

for i, c in enumerate(df_asr_summary.columns):
    cell = ws_asr_summary.cell(row=1, column=i+1)
    cell.value = c
    cell.font = font_colmun
    cell.alignment  = align_column
for index, row in df_asr_summary.iterrows():
    for j, c in enumerate(row):
        cell = ws_asr_summary.cell(row=index+2, column=j+1)
        cell.value = c
        cell.number_format = u"#,###,,"
        cell.font = font_cell
        cell.alignment  = align_cell
        
wb.save(xbrl_common.XBRL_ROOT_PATH + "/" + xbrl_common.ASR_ANALYSIS_FILE_NAME)
'''
with pd.ExcelWriter(xbrl_common.XBRL_ROOT_PATH + "/" + xbrl_common.ASR_ANALYSIS_FILE_NAME) as writer:
    df_asr_summary.to_excel(writer, sheet_name="全提出社")
    df_asr_summary_focused.to_excel(writer, sheet_name="Focused")
'''
print("  -> 完了")
'''
#---------------------------------------
# 多変量解析
#---------------------------------------
from sklearn.decomposition import PCA
from sklearn.preprocessing import StandardScaler, MinMaxScaler, RobustScaler

df_temp = df_asr_summary_focused[["EDINETコード", "提出者名", "従業員数", "純資産", "総資産", "売上高", "純利益", "営業CF"]]
mat_pca_input_data = (df_asr_summary_focused[["従業員数(5年平均)", "売上高(5年平均)", "純資産(5年平均)", "総資産(5年平均)", "純利益(5年平均)", "営業CF(5年平均)", "投資CF(5年平均)", "現金(5年平均)"]]).values


# 正規化
sc = MinMaxScaler()
input_data_scaled = sc.fit_transform(mat_pca_input_data)
# 主成分分析
pca = PCA(n_components=mat_pca_input_data.shape[1])
pca.fit(input_data_scaled)
X = np.dot(input_data_scaled,pca.components_.T)
df_temp["主成分1"] = X[:,0]
df_temp["主成分2"] = X[:,1]

print(pca.explained_variance_ratio_)
print(pca.components_)
'''


